#!/usr/bin/env python3
"""
个人月度账单统计工具
支持微信(xlsx)、支付宝(csv)、中国银行(pdf/xlsx)账单导入，生成月度/趋势统计报告

用法:
  bill import <文件> [--password PWD]   # 导入单文件（自动识别来源）
  bill import <文件1> <文件2> ...       # 批量导入
  bill report [YYYY-MM]                # 终端报告
  bill report --html [YYYY-MM]         # 生成HTML报告并在浏览器打开
  bill trend                           # 近6月趋势
  bill rules                           # 查看分类规则
  bill rules add <分类> <关键词>        # 添加分类规则
  bill export-guide                    # 显示各平台导出指南
"""

import argparse
import csv
import json
import os
import re
import shutil
import sqlite3
import sys
import webbrowser
from collections import defaultdict
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from io import StringIO
from pathlib import Path

# ── Paths ──
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "bills.db"
RULES_PATH = BASE_DIR / "rules.json"

# ── Default Category Rules ──
DEFAULT_RULES = {
    "餐饮美食": [
        "餐厅", "饭店", "美食", "外卖", "美团", "饿了么", "肯德基", "麦当劳",
        "火锅", "烧烤", "奶茶", "咖啡", "蛋糕", "面包", "小吃", "食堂", "快餐",
        "面馆", "小面", "鲜果", "香八度", "零满多", "元气鲜果茶", "水果",
        "渝味", "盒马", "外国语大学", "西安外国语", "锦绣",
        "牛肉面", "鸡公煲", "饸络", "夹馍", "肉夹馍", "麦浪", "面店",
        "小笼包", "飞鸿", "luckin", "coffee", "瑞幸", "炒饭", "煎饼",
        "西北狼", "杂粮煎饼", "五常大米",
    ],
    "交通出行": [
        "滴滴", "打车", "出租车", "地铁", "公交", "火车", "高铁", "机票",
        "加油", "高速", "停车", "共享单车", "骑行", "高德信息", "城市通",
    ],
    "购物消费": [
        "淘宝", "京东", "拼多多", "超市", "商场", "便利店", "百货", "天猫",
        "唯品会", "恩菲数码", "低卡博士", "珍视明", "猫与小鼠", "悦易尚",
        "格物致品", "抖音电商", "商贸", "电子店", "兰恩诗", "闲鱼",
    ],
    "住房物业": ["房租", "物业", "水电", "煤气", "宽带", "维修"],
    "休闲娱乐": ["电影", "KTV", "游戏", "旅游", "景点", "酒店", "按摩", "健身", "Apple"],
    "医疗健康": ["医院", "药店", "诊所", "挂号", "体检", "医药", "眼镜", "大药房"],
    "教育学习": ["学费", "培训", "书本", "考试", "知书网络", "知书科技"],
    "科技消费": ["话费", "流量", "充值", "wifi", "WiFi", "电子wifi",
               "手机", "电脑", "笔记本", "平板", "耳机", "音箱", "显示器",
               "键盘", "鼠标", "数码", "小米", "华为", "苹果", "vivo", "oppo",
               "三星", "配件", "数据线", "充电器", "硬盘", "U盘", "路由器",
               "DeepSeek", "深度求索", "API", "token", "Token"],
    "日用百货": ["日用品", "洗衣", "理发", "快递", "生活家", "宜佳合", "便利",
               "顺丰", "物流"],
    "服饰美容": ["衣服", "鞋子", "化妆品", "护肤品"],
    "人情往来": ["红包", "礼物", "捐款", "转账备注", "微信转账", "小荷包"],
    "金融服务": ["理财", "利息", "股票", "基金", "提现", "快捷提现",
               "银联入账", "自助存款", "快捷退款"],
    "休闲娱乐": ["电影", "KTV", "游戏", "旅游", "景点", "酒店", "按摩", "健身", "Apple", "K米"],
    "商业服务": ["深度求索", "DeepSeek"],
    "其他": [],
}

# Alipay category mapping to our standard categories
ALIPAY_CAT_MAP = {
    "餐饮美食": "餐饮美食",
    "交通出行": "交通出行",
    "日用百货": "日用百货",
    "购物消费": "购物消费",
    "医疗健康": "医疗健康",
    "住房物业": "住房物业",
    "休闲娱乐": "休闲娱乐",
    "文化休闲": "休闲娱乐",
    "商业服务": "商业服务",
    "教育培训": "教育学习",
    "通讯网络": "通讯网络",
    "充值缴费": "通讯网络",
    "服饰美容": "服饰美容",
    "美容美发": "服饰美容",
    "金融服务": "金融服务",
    "投资理财": "金融服务",
    "账户存取": "金融服务",
    "转账红包": "人情往来",
    "收入": "金融服务",
    "人情往来": "人情往来",
    "数码电器": "科技消费",
    "通讯网络": "科技消费",
    "充值缴费": "科技消费",
    "家居家装": "住房物业",
    "生活服务": "日用百货",
    "退款": None,  # handled by tx_type
}


# ═══════════════════════════════════════════════════════════════
# Database
# ═══════════════════════════════════════════════════════════════

def get_db():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    db = sqlite3.connect(str(DB_PATH))
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    return db


def init_db():
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            amount REAL NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('expense','income','neutral')),
            category TEXT NOT NULL DEFAULT '其他',
            merchant TEXT DEFAULT '',
            product TEXT DEFAULT '',
            source TEXT NOT NULL CHECK(source IN ('wechat','alipay','boc')),
            payment_method TEXT DEFAULT '',
            tx_id TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            is_duplicate INTEGER DEFAULT 0,
            dup_of INTEGER DEFAULT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(date, amount, merchant, source)
        );
        CREATE INDEX IF NOT EXISTS idx_tx_date ON transactions(date);
        CREATE INDEX IF NOT EXISTS idx_tx_source ON transactions(source);
        CREATE INDEX IF NOT EXISTS idx_tx_category ON transactions(category);
    """)
    db.commit()
    db.close()


def insert_transactions(txs):
    """Insert transactions, skipping duplicates by tx_id+source."""
    db = get_db()
    added = 0
    skipped = 0
    for tx in txs:
        try:
            db.execute(
                """INSERT INTO transactions
                   (date, amount, type, category, merchant, product, source,
                    payment_method, tx_id, remark)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (tx["date"], tx["amount"], tx["type"], tx["category"],
                 tx.get("merchant", ""), tx.get("product", ""), tx["source"],
                 tx.get("payment_method", ""), tx.get("tx_id", ""),
                 tx.get("remark", ""))
            )
            added += 1
        except sqlite3.IntegrityError:
            skipped += 1
    db.commit()
    db.close()
    return added, skipped


def query_transactions(month=None, source=None, category=None,
                       tx_type=None, include_dup=False):
    """Query transactions with optional filters."""
    db = get_db()
    conditions = []
    params = []

    if not include_dup:
        conditions.append("is_duplicate = 0")
    if month:
        conditions.append("strftime('%Y-%m', date) = ?")
        params.append(month)
    if source:
        conditions.append("source = ?")
        params.append(source)
    if category:
        conditions.append("category = ?")
        params.append(category)
    if tx_type:
        conditions.append("type = ?")
        params.append(tx_type)

    where = " AND ".join(conditions) if conditions else "1=1"
    rows = db.execute(
        f"SELECT * FROM transactions WHERE {where} ORDER BY date DESC",
        params
    ).fetchall()
    db.close()
    return [dict(r) for r in rows]


def get_available_months():
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT strftime('%Y-%m', date) as month FROM transactions "
        "ORDER BY month"
    ).fetchall()
    db.close()
    return [r["month"] for r in rows]


def get_monthly_summary(month):
    """Return {total_expense, total_income, count_expense, count_income}."""
    db = get_db()
    row = db.execute(
        """SELECT
             COALESCE(SUM(CASE WHEN type='expense' THEN amount ELSE 0 END), 0) as total_expense,
             COALESCE(SUM(CASE WHEN type='income' THEN amount ELSE 0 END), 0) as total_income,
             COUNT(CASE WHEN type='expense' THEN 1 END) as count_expense,
             COUNT(CASE WHEN type='income' THEN 1 END) as count_income
           FROM transactions
           WHERE strftime('%Y-%m', date) = ? AND is_duplicate = 0""",
        (month,)
    ).fetchone()
    db.close()
    return dict(row)


def get_category_breakdown(month):
    db = get_db()
    rows = db.execute(
        """SELECT category,
                 SUM(CASE WHEN type='expense' THEN amount ELSE 0 END) as expense,
                 SUM(CASE WHEN type='income' THEN amount ELSE 0 END) as income
           FROM transactions
           WHERE strftime('%Y-%m', date) = ? AND is_duplicate = 0
           GROUP BY category
           ORDER BY expense DESC""",
        (month,)
    ).fetchall()
    db.close()
    return [dict(r) for r in rows]


def get_daily_trend(month):
    db = get_db()
    rows = db.execute(
        """SELECT date, type, SUM(amount) as total
           FROM transactions
           WHERE strftime('%Y-%m', date) = ? AND is_duplicate = 0 AND type != 'neutral'
           GROUP BY date, type
           ORDER BY date""",
        (month,)
    ).fetchall()
    db.close()
    return [dict(r) for r in rows]


def get_monthly_trend(months=6):
    """Get expense totals by category for the last N months."""
    db = get_db()
    rows = db.execute(
        """SELECT strftime('%Y-%m', date) as month,
                 category,
                 SUM(CASE WHEN type='expense' THEN amount ELSE 0 END) as expense
           FROM transactions
           WHERE is_duplicate = 0 AND type = 'expense'
           GROUP BY month, category
           ORDER BY month, expense DESC"""
    ).fetchall()
    db.close()
    return [dict(r) for r in rows]


def mark_duplicates(dup_pairs):
    """Mark transactions as duplicates. dup_pairs is list of (dup_id, original_id)."""
    if not dup_pairs:
        return 0
    db = get_db()
    count = 0
    for dup_id, orig_id in dup_pairs:
        db.execute(
            "UPDATE transactions SET is_duplicate=1, dup_of=? WHERE id=?",
            (orig_id, dup_id)
        )
        count += 1
    db.commit()
    db.close()
    return count


# ═══════════════════════════════════════════════════════════════
# Category Engine
# ═══════════════════════════════════════════════════════════════

def load_rules():
    if RULES_PATH.exists():
        with open(RULES_PATH, "r") as f:
            return json.load(f)
    # Auto-save defaults on first run
    rules = dict(DEFAULT_RULES)
    save_rules(rules)
    return rules


def save_rules(rules):
    with open(RULES_PATH, "w") as f:
        json.dump(rules, f, ensure_ascii=False, indent=2)


def categorize(merchant, product, tx_type, source, alipay_category=""):
    """
    Classify a transaction based on merchant name and product description.
    Returns category name string.
    Keyword matching takes precedence over Alipay's built-in category,
    so user rules can override Alipay's classification.
    """
    rules = load_rules()
    text = f"{merchant} {product}"

    # Keyword match first (user rules override Alipay)
    for category, keywords in rules.items():
        if category == "其他":
            continue
        for kw in keywords:
            if kw in text:
                return category

    # Alipay自带分类作为兜底
    if source == "alipay" and alipay_category:
        alipay_cat = alipay_category.strip()
        if alipay_cat in ALIPAY_CAT_MAP:
            mapped = ALIPAY_CAT_MAP[alipay_cat]
            if mapped is not None:
                return mapped

    return "其他"


def recategorize_all():
    """Re-categorize all transactions based on current rules."""
    db = get_db()
    txs = db.execute("SELECT * FROM transactions").fetchall()
    updated = 0
    for tx in txs:
        txd = dict(tx)
        alipay_cat = ""
        remark = txd.get("remark", "")
        # Extract stored Alipay category from remark
        if remark.startswith("[ALIPAY_CAT:"):
            parts = remark.split("]", 1)
            alipay_cat = parts[0][12:]  # remove "[ALIPAY_CAT:"
            remark = parts[1] if len(parts) > 1 else ""
        new_cat = categorize(
            txd.get("merchant", ""),
            txd.get("product", ""),
            txd["type"],
            txd["source"],
            alipay_category=alipay_cat
        )
        if new_cat != tx["category"]:
            db.execute("UPDATE transactions SET category=? WHERE id=?",
                       (new_cat, tx["id"]))
            updated += 1
    db.commit()
    db.close()
    return updated


# ═══════════════════════════════════════════════════════════════
# Parsers
# ═══════════════════════════════════════════════════════════════

def excel_serial_to_date(serial):
    """Convert Excel serial number to datetime string."""
    try:
        serial = float(serial)
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=serial)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError):
        return str(serial)


def parse_amount(val):
    """Parse amount string/number, removing commas and currency symbols."""
    if isinstance(val, (int, float)):
        return float(val)
    if not val:
        return 0.0
    val = str(val).replace(",", "").replace("¥", "").replace("￥", "").strip()
    try:
        return float(val)
    except ValueError:
        return 0.0


def detect_format(filepath):
    """Detect source format: 'wechat', 'alipay', 'boc', or None."""
    name = filepath.name.lower()
    path_str = str(filepath).lower()

    if "wechat" in path_str or "微信" in path_str:
        if name.endswith(".xlsx"):
            return "wechat"
    if "alipay" in path_str or "支付宝" in path_str:
        if name.endswith(".csv"):
            return "alipay"
    if "boc" in path_str or "中国银行" in path_str or "中国银行" in str(filepath.parent).lower():
        if name.endswith(".pdf"):
            return "boc_pdf"
        if name.endswith(".xlsx"):
            return "boc_xlsx"

    # Fallback: try by extension
    if name.endswith(".xlsx"):
        return "wechat"  # most likely
    if name.endswith(".csv"):
        return "alipay"
    if name.endswith(".pdf"):
        return "boc_pdf"

    return None


# ── WeChat Parser ──

def parse_wechat(filepath):
    """Parse WeChat xlsx export file."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Find header row
    header_row = None
    data_start = None
    for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True)):
        if row[0] and "交易时间" in str(row[0]):
            header_row = [str(c) if c else "" for c in row]
            data_start = i + 1
            break

    if header_row is None:
        return [], "无法找到微信账单表头"

    # Map column indices
    col_map = {}
    for idx, col_name in enumerate(header_row):
        col_map[col_name.strip()] = idx

    txs = []
    for row in ws.iter_rows(min_row=data_start + 1, values_only=True):
        try:
            date_val = row[col_map.get("交易时间", 0)]
            if not date_val:
                continue

            date_str = excel_serial_to_date(date_val)
            merchant = str(row[col_map.get("交易对方", 2)] or "")
            product = str(row[col_map.get("商品", 3)] or "")
            tx_type_str = str(row[col_map.get("收/支", 4)] or "")
            amount = parse_amount(row[col_map.get("金额(元)", 5)])
            payment_method = str(row[col_map.get("支付方式", 6)] or "")
            tx_id = str(row[col_map.get("交易单号", 8)] or "")
            remark = str(row[col_map.get("备注", 10)] or "")

            if tx_type_str == "支出":
                tx_type = "expense"
            elif tx_type_str == "收入":
                tx_type = "income"
            else:
                tx_type = "neutral"

            category = categorize(merchant, product, tx_type, "wechat")

            txs.append({
                "date": date_str[:10],
                "amount": amount,
                "type": tx_type,
                "category": category,
                "merchant": merchant,
                "product": product,
                "source": "wechat",
                "payment_method": payment_method,
                "tx_id": tx_id,
                "remark": remark,
            })
        except Exception:
            continue

    wb.close()
    return txs, None


# ── Alipay Parser ──

def parse_alipay(filepath):
    """Parse Alipay CSV export file."""
    # Try encoding detection
    encodings = ["gbk", "gb2312", "utf-8", "utf-8-sig"]
    text = None
    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc) as f:
                text = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if text is None:
        return [], "无法识别支付宝账单编码"

    # Find header line
    lines = text.split("\n")
    header_idx = None
    for i, line in enumerate(lines):
        if "交易时间" in line and "交易对方" in line:
            header_idx = i
            break

    if header_idx is None:
        return [], "无法找到支付宝账单表头"

    reader = csv.DictReader(StringIO("\n".join(lines[header_idx:])))
    txs = []

    for row in reader:
        try:
            date_str = (row.get("交易时间", "") or "").strip()
            alipay_cat = (row.get("交易分类", "") or "").strip()
            merchant = (row.get("交易对方", "") or "").strip()
            product = (row.get("商品说明", "") or "").strip()
            tx_type_str = (row.get("收/支", "") or "").strip()
            amount = parse_amount(row.get("金额", "0"))
            payment_method = (row.get("收/付款方式", "") or "").strip()
            tx_id = (row.get("交易订单号", "") or "").strip()
            remark = (row.get("备注", "") or "").strip()

            if not date_str:
                continue

            if tx_type_str in ("支出", "已支出"):
                tx_type = "expense"
            elif tx_type_str in ("收入", "已收入"):
                tx_type = "income"
            else:
                tx_type = "neutral"

            category = categorize(merchant, product, tx_type, "alipay",
                                  alipay_category=alipay_cat)

            # Store Alipay category in remark for later recategorization
            stored_remark = remark
            if alipay_cat:
                stored_remark = f"[ALIPAY_CAT:{alipay_cat}]{remark}"

            txs.append({
                "date": date_str[:10],
                "amount": amount,
                "type": tx_type,
                "category": category,
                "merchant": merchant,
                "product": product,
                "source": "alipay",
                "payment_method": payment_method,
                "tx_id": tx_id,
                "remark": stored_remark,
            })
        except Exception:
            continue

    return txs, None


# ── BOC PDF Parser ──

def parse_boc_pdf(filepath, password=None):
    """Parse Bank of China PDF statement."""
    try:
        import pdfplumber
    except ImportError:
        return [], "需要安装 pdfplumber: pip install pdfplumber"

    try:
        pdf = pdfplumber.open(str(filepath), password=password)
    except Exception as e:
        return [], f"无法打开PDF（可能需要密码）: {e}"

    txs = []
    for page in pdf.pages:
        tables = page.extract_tables()
        for table in tables:
            for row in table:
                if not row or not row[0]:
                    continue
                # Skip header rows
                if "记账日期" in str(row[0]):
                    continue

                try:
                    date_str = str(row[0] or "").strip().replace("\n", "")
                    # time_str = str(row[1] or "").strip().replace("\n", "")
                    amount_str = str(row[3] or "").strip().replace("\n", "")
                    amount_str = amount_str.replace(",", "")
                    merchant_raw = str(row[9] or "").strip().replace("\n", "")
                    tx_name = str(row[5] or "").strip().replace("\n", "")
                    counterparty = str(row[10] or "").strip().replace("\n", "")
                    bank_name = str(row[11] or "").strip().replace("\n", "")

                    if not date_str or not amount_str:
                        continue

                    amount = float(amount_str)

                    if amount < 0:
                        tx_type = "expense"
                        amount = abs(amount)
                    else:
                        tx_type = "income"

                    merchant = merchant_raw
                    product = tx_name

                    category = categorize(merchant, product, tx_type, "boc")

                    txs.append({
                        "date": date_str,
                        "amount": amount,
                        "type": tx_type,
                        "category": category,
                        "merchant": merchant,
                        "product": product,
                        "source": "boc",
                        "payment_method": tx_name,
                        "tx_id": "",
                        "remark": f"{counterparty} {bank_name}",
                    })
                except (ValueError, IndexError):
                    continue

    pdf.close()
    return txs, None


# ── BOC XLSX Parser (fallback for pre-converted files) ──

def parse_boc_xlsx(filepath):
    """Parse BOC xlsx (converted from PDF)."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath)
    txs = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_found = False
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            if "记账日期" in str(row[0]):
                header_found = True
                continue
            if not header_found:
                continue

            try:
                date_val = row[0]
                if isinstance(date_val, (int, float)):
                    date_str = excel_serial_to_date(date_val)[:10]
                else:
                    date_str = str(date_val).strip()

                amount = parse_amount(row[3])

                if amount < 0:
                    tx_type = "expense"
                    amount = abs(amount)
                else:
                    tx_type = "income"

                merchant = str(row[9] or "").strip().replace("\n", "") if len(row) > 9 else ""
                product = str(row[5] or "").strip().replace("\n", "") if len(row) > 5 else ""

                category = categorize(merchant, product, tx_type, "boc")

                txs.append({
                    "date": date_str,
                    "amount": amount,
                    "type": tx_type,
                    "category": category,
                    "merchant": merchant,
                    "product": product,
                    "source": "boc",
                    "payment_method": product,
                    "tx_id": "",
                    "remark": "",
                })
            except (ValueError, IndexError):
                continue

    wb.close()
    return txs, None


def parse_file(filepath, password=None):
    """Parse a single file, auto-detecting format. Returns (txs, error)."""
    fmt = detect_format(filepath)

    if fmt == "wechat":
        return parse_wechat(filepath)
    elif fmt == "alipay":
        return parse_alipay(filepath)
    elif fmt == "boc_pdf":
        return parse_boc_pdf(filepath, password)
    elif fmt == "boc_xlsx":
        return parse_boc_xlsx(filepath)
    else:
        # Try extensions as last resort
        if filepath.suffix.lower() == ".csv":
            return parse_alipay(filepath)
        elif filepath.suffix.lower() == ".xlsx":
            return parse_wechat(filepath)
        elif filepath.suffix.lower() == ".pdf":
            return parse_boc_pdf(filepath, password)
        else:
            return [], f"无法识别文件格式: {filepath}"


# ═══════════════════════════════════════════════════════════════
# Deduplication
# ═══════════════════════════════════════════════════════════════

def check_source_overlap(txs):
    """Check if parsed transactions overlap with existing data in DB."""
    if not txs:
        return {}
    sources = set(t["source"] for t in txs)
    overlaps = {}
    db = get_db()
    for src in sources:
        src_txs = [t for t in txs if t["source"] == src]
        if not src_txs:
            continue
        min_d = min(t["date"] for t in src_txs)
        max_d = max(t["date"] for t in src_txs)
        existing = db.execute(
            "SELECT COUNT(*) as cnt FROM transactions WHERE source=? AND date BETWEEN ? AND ?",
            (src, min_d, max_d)
        ).fetchone()
        if existing["cnt"] > 0:
            overlaps[src] = {"count": existing["cnt"], "min": min_d, "max": max_d}
    db.close()
    return overlaps


def archive_import(filepath, source, txs):
    """Copy imported file to data/imports/ with standardized name."""
    archive_dir = DATA_DIR / "imports"
    archive_dir.mkdir(parents=True, exist_ok=True)
    src_txs = [t for t in txs if t["source"] == source]
    if not src_txs:
        return None
    min_d = min(t["date"] for t in src_txs).replace("-", "")
    max_d = max(t["date"] for t in src_txs).replace("-", "")
    ext = filepath.suffix.lower()
    new_name = f"{source}_{min_d}_{max_d}{ext}"
    dest = archive_dir / new_name
    shutil.copy2(filepath, dest)
    return dest


def dedup_same_source():
    """Deduplicate within same source by date + amount + fuzzy merchant match."""
    db = get_db()
    txs = db.execute(
        "SELECT * FROM transactions WHERE is_duplicate=0"
    ).fetchall()
    db.close()

    groups = defaultdict(list)
    for tx in txs:
        txd = dict(tx)
        key = (txd["source"], txd["date"], round(txd["amount"], 2))
        groups[key].append(txd)

    dup_pairs = []
    for key, group in groups.items():
        if len(group) < 2:
            continue
        for i in range(len(group)):
            for j in range(i + 1, len(group)):
                m1 = (group[i].get("merchant", "") + group[i].get("product", "")).lower()
                m2 = (group[j].get("merchant", "") + group[j].get("product", "")).lower()
                if not m1 or not m2:
                    continue
                similarity = SequenceMatcher(None, m1, m2).ratio()
                if similarity > 0.5:
                    if group[i]["id"] < group[j]["id"]:
                        dup_pairs.append((group[j]["id"], group[i]["id"]))
                    else:
                        dup_pairs.append((group[i]["id"], group[j]["id"]))

    return mark_duplicates(dup_pairs)


def find_and_mark_duplicates():
    """
    Find cross-source duplicates.
    Rule: same date, same amount (±0.01), different sources → mark BOC as dup.
    Priority: alipay > wechat > boc
    """
    db = get_db()
    txs = db.execute(
        "SELECT * FROM transactions WHERE is_duplicate=0 AND type='expense'"
    ).fetchall()
    db.close()

    # Group by (date, amount_rounded)
    groups = defaultdict(list)
    for tx in txs:
        tx = dict(tx)
        key = (tx["date"], round(tx["amount"], 2))
        groups[key].append(tx)

    dup_pairs = []
    source_priority = {"alipay": 0, "wechat": 1, "boc": 2}

    for key, group in groups.items():
        if len(group) < 2:
            continue
        group.sort(key=lambda x: source_priority.get(x["source"], 99))
        # Mark lower-priority ones as duplicates of the highest-priority
        best = group[0]
        for tx in group[1:]:
            if tx["source"] != best["source"]:
                dup_pairs.append((tx["id"], best["id"]))

    count = mark_duplicates(dup_pairs)
    return count


# ═══════════════════════════════════════════════════════════════
# Terminal Report
# ═══════════════════════════════════════════════════════════════

def terminal_report(month):
    """Print a formatted terminal report for the given month."""
    summary = get_monthly_summary(month)
    categories = get_category_breakdown(month)

    total_expense = summary["total_expense"]
    total_income = summary["total_income"]

    # Month-over-month comparison
    prev_month = previous_month(month)
    prev_summary = get_monthly_summary(prev_month) if prev_month else None

    print()
    print("╔" + "═" * 54 + "╗")
    print(f"║  {month} 月度账单报告".ljust(47) + "║")
    print("╠" + "═" * 54 + "╣")

    # Summary line
    expense_str = f"总支出: ¥{total_expense:,.2f}"
    income_str = f"总收入: ¥{total_income:,.2f}"
    balance = total_income - total_expense
    balance_str = f"结余: ¥{balance:+,.2f}"
    count_str = f"笔数: {summary['count_expense']}支出 {summary['count_income']}收入"
    print(f"║  {expense_str:<20} {income_str:<22} ║")
    print(f"║  {balance_str:<20} {count_str:<22} ║")

    # MoM comparison
    if prev_summary and prev_summary["total_expense"] > 0:
        prev_exp = prev_summary["total_expense"]
        change = (total_expense - prev_exp) / prev_exp * 100
        arrow = "↑" if change > 0 else "↓"
        mom_str = f"环比({prev_month}): {arrow} {abs(change):.1f}%"
        print(f"║  {mom_str:<44} ║")

    print("╠" + "═" * 54 + "╣")
    print("║  分类支出明细:".ljust(47) + "║")

    if total_expense == 0:
        print("║  (本月无支出记录)".ljust(47) + "║")
    else:
        max_bar_width = 24
        for cat in categories:
            if cat["expense"] <= 0:
                continue
            pct = cat["expense"] / total_expense * 100
            bar_len = int(pct / 100 * max_bar_width)
            bar = "█" * bar_len
            line = f"  {cat['category']:<8} {bar} {pct:5.1f}% ¥{cat['expense']:,.2f}"
            # Pad to ensure at least 3 spaces at end
            print(f"║{line:<54}║")

    print("╚" + "═" * 54 + "╝")
    print()


def terminal_trend(months=6):
    """Print 6-month expense trend by category."""
    all_months = get_available_months()
    if not all_months:
        print("暂无数据")
        return

    recent = all_months[-months:]
    trend_data = get_monthly_trend(months)

    # Aggregate by month and category
    month_cat = defaultdict(lambda: defaultdict(float))
    all_cats = set()
    for row in trend_data:
        if row["month"] in recent:
            month_cat[row["month"]][row["category"]] += row["expense"]
            all_cats.add(row["category"])

    # Get top categories by total
    cat_totals = defaultdict(float)
    for month, cats in month_cat.items():
        for cat, val in cats.items():
            cat_totals[cat] += val
    top_cats = sorted(cat_totals, key=cat_totals.get, reverse=True)[:6]

    print()
    print("═" * 60)
    print("  近{}月支出趋势（分类别）".format(len(recent)))
    print("═" * 60)

    # Header
    header = f"{'分类':<10}"
    for m in recent:
        header += f"{m[5:]:>8}"
    print(header)
    print("-" * (10 + 8 * len(recent)))

    # Rows
    for cat in top_cats:
        line = f"{cat:<10}"
        for m in recent:
            val = month_cat[m].get(cat, 0)
            if val > 0:
                line += f"¥{val:>7.0f}"
            else:
                line += f"{'':>8}"
        print(line)

    # Total row
    print("-" * (10 + 8 * len(recent)))
    total_line = f"{'合计':<10}"
    for m in recent:
        total = sum(month_cat[m].values())
        total_line += f"¥{total:>7.0f}"
    print(total_line)
    print()


# ═══════════════════════════════════════════════════════════════
# HTML Report
# ═══════════════════════════════════════════════════════════════

def html_report(month=None, open_browser=True):
    """Generate HTML report and optionally open in browser."""
    if month is None:
        available = get_available_months()
        if not available:
            print("暂无数据，请先导入账单")
            return None
        month = available[-1]

    summary = get_monthly_summary(month)
    categories = get_category_breakdown(month)
    daily = get_daily_trend(month)
    all_months = get_available_months()
    trend_data = get_monthly_trend(len(all_months))

    total_expense = summary["total_expense"]
    total_income = summary["total_income"]
    balance = total_income - total_expense

    # Previous month comparison
    prev_month = previous_month(month)
    prev_summary = get_monthly_summary(prev_month) if prev_month else None
    mom_change = 0
    if prev_summary and prev_summary["total_expense"] > 0:
        mom_change = (total_expense - prev_summary["total_expense"]) / prev_summary["total_expense"] * 100

    # Category data for pie chart
    cat_labels = []
    cat_values = []
    cat_colors = [
        "#FF6384", "#36A2EB", "#FFCE56", "#4BC0C0", "#9966FF",
        "#FF9F40", "#7BC8A4", "#E8A87C", "#85C1E9", "#F1948A",
        "#BB8FCE", "#85C1E9", "#F7DC6F", "#82E0AA",
    ]
    for cat in categories:
        if cat["expense"] > 0:
            cat_labels.append(cat["category"])
            cat_values.append(round(cat["expense"], 2))

    # Daily trend data (total)
    days = sorted(set(d["date"] for d in daily))
    daily_expense = []
    daily_income = []
    for day in days:
        exp = sum(d["total"] for d in daily if d["date"] == day and d["type"] == "expense")
        inc = sum(d["total"] for d in daily if d["date"] == day and d["type"] == "income")
        daily_expense.append(round(exp, 2))
        daily_income.append(round(inc, 2))

    # Daily by category data
    db2 = get_db()
    daily_cat_rows = db2.execute(
        """SELECT date, category, SUM(amount) as total
           FROM transactions
           WHERE strftime('%Y-%m', date) = ? AND type='expense' AND is_duplicate=0
           GROUP BY date, category ORDER BY date""",
        (month,)
    ).fetchall()
    db2.close()

    # Build daily category datasets (top N categories by total)
    day_set = sorted(set(r["date"] for r in daily_cat_rows))
    cat_daily_totals = defaultdict(float)
    for r in daily_cat_rows:
        cat_daily_totals[r["category"]] += r["total"]
    top_daily_cats = sorted(cat_daily_totals, key=cat_daily_totals.get, reverse=True)[:5]

    daily_cat_datasets = []
    for cat in top_daily_cats:
        cat_day_map = {r["date"]: r["total"] for r in daily_cat_rows if r["category"] == cat}
        vals = [round(cat_day_map.get(d, 0), 2) for d in day_set]
        daily_cat_datasets.append({"label": cat, "data": vals})

    # Monthly trend data (top 5 categories + 其他)
    recent_months = all_months[-12:]
    month_cat_data = defaultdict(lambda: defaultdict(float))
    all_cats_trend = set()
    for row in trend_data:
        if row["month"] in recent_months:
            month_cat_data[row["month"]][row["category"]] += row["expense"]
            all_cats_trend.add(row["category"])

    # Top categories by overall total
    cat_totals = defaultdict(float)
    for m, cats in month_cat_data.items():
        for c, v in cats.items():
            cat_totals[c] += v
    top_trend_cats = sorted(cat_totals, key=cat_totals.get, reverse=True)[:6]

    trend_datasets = []
    for cat in top_trend_cats:
        vals = [round(month_cat_data[m].get(cat, 0), 2) for m in recent_months]
        trend_datasets.append({"label": cat, "values": vals})

    # Previous month category data for comparison
    prev_categories = {}
    if prev_month:
        prev_cats = get_category_breakdown(prev_month)
        prev_categories = {c["category"]: c["expense"] for c in prev_cats}

    # Build当月vs上月 category comparison data
    compare_cats = []
    compare_curr = []
    compare_prev = []
    for cat in categories:
        if cat["expense"] > 0:
            compare_cats.append(cat["category"])
            compare_curr.append(round(cat["expense"], 2))
            compare_prev.append(round(prev_categories.get(cat["category"], 0), 2))

    # Build trend chart datasets with colors in Python
    trend_datasets_json = []
    for i, d in enumerate(trend_datasets):
        color = cat_colors[i % len(cat_colors)]
        trend_datasets_json.append({
            "label": d["label"],
            "data": d["values"],
            "borderColor": color,
            "backgroundColor": color + "20",
            "fill": False,
            "tension": 0.3,
        })

    # Category counts
    db = get_db()
    cat_counts = {}
    count_rows = db.execute(
        """SELECT category, COUNT(*) as cnt FROM transactions
           WHERE strftime('%Y-%m', date) = ? AND type='expense' AND is_duplicate=0
           GROUP BY category""", (month,)
    ).fetchall()
    cat_counts = {r["category"]: r["cnt"] for r in count_rows}
    db.close()

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>账单报告 - {month}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root {{ --bg:#f5f6fa; --card:#fff; --text:#2d3436; --muted:#636e72;
       --expense:#e74c3c; --income:#27ae60; --primary:#0984e3; --border:#dfe6e9; }}
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','PingFang SC','Microsoft YaHei',sans-serif;
       background:var(--bg); color:var(--text); padding:24px; }}
h1 {{ font-size:24px; margin-bottom:20px; }}
.cards {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(200px,1fr)); gap:16px; margin-bottom:24px; }}
.card {{ background:var(--card); border-radius:12px; padding:20px; box-shadow:0 1px 3px rgba(0,0,0,.08); }}
.card .label {{ font-size:13px; color:var(--muted); margin-bottom:6px; }}
.card .value {{ font-size:28px; font-weight:700; }}
.card.expense .value {{ color:var(--expense); }}
.card.income .value {{ color:var(--income); }}
.card.balance .value {{ color:var(--primary); }}
.card .sub {{ font-size:12px; color:var(--muted); margin-top:4px; }}
.charts {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; margin-bottom:24px; }}
@media(max-width:768px){{ .charts{{grid-template-columns:1fr;}} }}
.chart-box {{ background:var(--card); border-radius:12px; padding:20px; box-shadow:0 1px 3px rgba(0,0,0,.08); }}
.chart-box h3 {{ font-size:15px; margin-bottom:12px; color:var(--muted); }}
.chart-box canvas {{ max-height:300px; }}
.trend-box {{ background:var(--card); border-radius:12px; padding:20px; box-shadow:0 1px 3px rgba(0,0,0,.08); margin-bottom:24px; }}
.trend-box h3 {{ font-size:15px; margin-bottom:12px; color:var(--muted); }}
.trend-box canvas {{ max-height:320px; }}
table {{ width:100%; border-collapse:collapse; background:var(--card); border-radius:12px;
        overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,.08); }}
th,td {{ padding:12px 16px; text-align:left; }}
th {{ background:#f8f9fa; font-size:12px; color:var(--muted); text-transform:uppercase; }}
td {{ border-top:1px solid var(--border); font-size:14px; }}
tr:hover td {{ background:#f8f9fa; }}
.bar-cell {{ display:flex; align-items:center; gap:8px; }}
.bar {{ height:8px; border-radius:4px; background:var(--primary); }}
.number {{ text-align:right; font-variant-numeric:tabular-nums; }}
</style>
</head>
<body>
<h1>{month} 月度账单报告</h1>

<div class="cards">
  <div class="card expense">
    <div class="label">总支出</div>
    <div class="value">¥{total_expense:,.2f}</div>
    <div class="sub">{summary['count_expense']} 笔</div>
  </div>
  <div class="card income">
    <div class="label">总收入</div>
    <div class="value">¥{total_income:,.2f}</div>
    <div class="sub">{summary['count_income']} 笔</div>
  </div>
  <div class="card balance">
    <div class="label">结余</div>
    <div class="value">¥{balance:+,.2f}</div>
    <div class="sub">环比 {prev_month}: {mom_change:+.1f}%</div>
  </div>
</div>

<div class="charts">
  <div class="chart-box">
    <h3>分类支出占比</h3>
    <canvas id="pieChart"></canvas>
  </div>
  <div class="chart-box">
    <h3>当月分类支出对比 (vs {prev_month})</h3>
    <canvas id="categoryCompareChart"></canvas>
  </div>
</div>

<div class="trend-box">
  <h3>当月每日收支趋势</h3>
  <canvas id="dailyChart"></canvas>
</div>

<div class="trend-box">
  <h3>当月每日分类支出趋势</h3>
  <canvas id="dailyCatChart"></canvas>
</div>

<div class="trend-box">
  <h3>近12月分类趋势</h3>
  <canvas id="trendChart"></canvas>
</div>

<table>
  <thead>
    <tr><th>分类</th><th class="number">支出金额</th><th class="number">占比</th><th class="number">笔数</th><th class="number">环比</th></tr>
  </thead>
  <tbody>
"""

    for cat in categories:
        if cat["expense"] > 0:
            pct = cat["expense"] / total_expense * 100 if total_expense > 0 else 0
            cnt = cat_counts.get(cat["category"], 0)
            prev_exp = prev_categories.get(cat["category"], 0)
            if prev_exp > 0:
                mom = (cat["expense"] - prev_exp) / prev_exp * 100
                mom_str = f"{mom:+.1f}%"
            else:
                mom_str = "-"
            html += f"""
    <tr>
      <td>{cat['category']}</td>
      <td class="number">¥{cat['expense']:,.2f}</td>
      <td class="number">
        <div class="bar-cell">
          <div class="bar" style="width:{max(pct, 2)}px"></div>
          {pct:.1f}%
        </div>
      </td>
      <td class="number">{cnt}</td>
      <td class="number">{mom_str}</td>
    </tr>"""

    html += f"""
  </tbody>
</table>

<h2 style="margin-top:24px;">当月每笔支出（高到低）</h2>
<table>
  <thead>
    <tr><th>日期</th><th>商户</th><th>商品</th><th>分类</th><th class="number">金额</th></tr>
  </thead>
  <tbody>
"""

    # Query all expense transactions for this month
    db3 = get_db()
    all_txs = db3.execute(
        """SELECT date, merchant, product, category, amount FROM transactions
           WHERE strftime('%Y-%m', date) = ? AND type='expense' AND is_duplicate=0
           ORDER BY amount DESC""",
        (month,)
    ).fetchall()
    db3.close()

    for tx in all_txs:
        html += f"""
    <tr>
      <td>{tx['date']}</td>
      <td>{tx['merchant'][:30]}</td>
      <td style="color:var(--muted);font-size:13px;">{tx['product'][:30]}</td>
      <td>{tx['category']}</td>
      <td class="number">¥{tx['amount']:,.2f}</td>
    </tr>"""

    html += f"""
  </tbody>
</table>

<script>
const COLORS = {json.dumps(cat_colors)};

// Pie chart
new Chart(document.getElementById('pieChart'), {{
  type: 'doughnut',
  data: {{
    labels: {json.dumps(cat_labels)},
    datasets: [{{
      data: {json.dumps(cat_values)},
      backgroundColor: COLORS.slice(0, {len(cat_labels)}),
      borderWidth: 0,
    }}]
  }},
  options: {{
    responsive: true,
    plugins: {{ legend: {{ position: 'right', labels: {{ padding: 16, usePointStyle: true }} }} }},
  }}
}});

// Category comparison chart (当月 vs 上月)
new Chart(document.getElementById('categoryCompareChart'), {{
  type: 'bar',
  data: {{
    labels: {json.dumps(compare_cats)},
    datasets: [
      {{
        label: '{month}',
        data: {json.dumps(compare_curr)},
        backgroundColor: '#0984e3',
        borderRadius: 3,
      }},
      {{
        label: '{prev_month}',
        data: {json.dumps(compare_prev)},
        backgroundColor: '#b2bec3',
        borderRadius: 3,
      }}
    ]
  }},
  options: {{
    responsive: true,
    plugins: {{ legend: {{ position: 'top' }} }},
    scales: {{
      y: {{ beginAtZero: true, ticks: {{ callback: function(v) {{ return '¥' + v; }} }} }}
    }},
  }}
}});

// Daily chart
new Chart(document.getElementById('dailyChart'), {{
  type: 'bar',
  data: {{
    labels: {json.dumps([d[-5:] for d in days])},
    datasets: [
      {{
        label: '支出',
        data: {json.dumps(daily_expense)},
        backgroundColor: '#e74c3c',
        borderRadius: 2,
      }},
      {{
        label: '收入',
        data: {json.dumps(daily_income)},
        backgroundColor: '#27ae60',
        borderRadius: 2,
      }}
    ]
  }},
  options: {{
    responsive: true,
    plugins: {{ legend: {{ position: 'top' }} }},
    scales: {{ y: {{ beginAtZero: true }} }},
  }}
}});

// Daily by category chart
new Chart(document.getElementById('dailyCatChart'), {{
  type: 'line',
  data: {{
    labels: {json.dumps([d[-5:] for d in day_set])},
    datasets: {json.dumps([{**d, 'borderColor': cat_colors[i % len(cat_colors)], 'backgroundColor': cat_colors[i % len(cat_colors)] + '20', 'fill': False, 'tension': 0.3, 'pointRadius': 3} for i, d in enumerate(daily_cat_datasets)])},
  }},
  options: {{
    responsive: true,
    plugins: {{ legend: {{ position: 'top' }} }},
    scales: {{ y: {{ beginAtZero: true, ticks: {{ callback: function(v){{return \"¥\"+v}} }} }} }},
    interaction: {{ intersect: false, mode: 'index' }},
  }}
}});

// Trend chart
new Chart(document.getElementById('trendChart'), {{
  type: 'line',
  data: {{
    labels: {json.dumps([m[-5:] for m in recent_months])},
    datasets: {json.dumps(trend_datasets_json)},
  }},
  options: {{
    responsive: true,
    plugins: {{ legend: {{ position: 'top' }} }},
    scales: {{ y: {{ beginAtZero: true }} }},
    interaction: {{ intersect: false, mode: 'index' }},
  }}
}});
</script>
</body>
</html>"""

    report_path = DATA_DIR / f"report_{month}.html"
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(html)

    if open_browser:
        webbrowser.open(f"file://{report_path}")

    return report_path


# ═══════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════

def previous_month(month):
    """Return the previous month string like '2026-04'."""
    try:
        y, m = month.split("-")
        dt = datetime(int(y), int(m), 1) - timedelta(days=1)
        return dt.strftime("%Y-%m")
    except Exception:
        return None


def show_export_guide():
    print("""
╔══════════════════════════════════════════════════════════════╗
║                     账单导出指南                             ║
╠══════════════════════════════════════════════════════════════╣
║                                                              ║
║  [微信支付]                                                  ║
║  我 → 服务 → 钱包 → 账单 → 常见问题 → 下载账单               ║
║  → 选择"用于个人对账" → 输入邮箱 → 查收邮件下载 xlsx          ║
║                                                              ║
║  [支付宝]                                                    ║
║  我的 → 账单 → ... → 开具交易流水证明                        ║
║  → 选择"用于个人对账" → 输入邮箱 → 查收邮件下载 csv           ║
║                                                              ║
║  [中国银行]                                                  ║
║  手机银行App → 交易查询 → 导出/打印 → 导出为 PDF              ║
║  (需要账户密码打开，每次密码不同)                              ║
║                                                              ║
║  导入命令:                                                   ║
║  bill import 微信.xlsx 支付宝.csv 中行.pdf --password xxx     ║
║                                                              ║
╚══════════════════════════════════════════════════════════════╝
""")


def show_rules():
    rules = load_rules()
    print("\n当前分类规则:\n")
    for cat, keywords in rules.items():
        print(f"  {cat}: {', '.join(keywords) if keywords else '(无关键词)'}")
    print()


def add_rule(category, keyword):
    rules = load_rules()
    if category not in rules:
        print(f"分类 '{category}' 不存在，可用分类: {', '.join(rules.keys())}")
        return
    if keyword in rules[category]:
        print(f"关键词 '{keyword}' 已存在于分类 '{category}'")
        return
    rules[category].append(keyword)
    save_rules(rules)
    print(f"已添加: {category} ← '{keyword}'")
    print(f"重新分类已有交易: {recategorize_all()} 条交易已更新")


# ═══════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════

def cmd_import(args):
    init_db()
    total_dup = 0

    for filepath in args.files:
        path = Path(filepath)
        if not path.exists():
            print(f"[ERROR] 文件不存在: {filepath}", file=sys.stderr)
            continue

        print(f"正在解析: {path.name} ... ", end="", flush=True)
        txs, error = parse_file(path, password=args.password)

        if error:
            print(f"[ERROR] {error}", file=sys.stderr)
            continue

        if not txs:
            print("[WARN] 未解析到交易记录", file=sys.stderr)
            continue

        # Overlap detection
        source = txs[0]["source"]
        overlaps = check_source_overlap(txs)
        if overlaps:
            for src, info in overlaps.items():
                print(f"\n  注意: {src} 在 {info['min']} ~ {info['max']} 区间已有 {info['count']} 条记录")
                print(f"  重复记录将被自动跳过（按 日期+金额+商户 去重）")

        added, skipped = insert_transactions(txs)
        expense_count = sum(1 for t in txs if t["type"] == "expense")
        income_count = sum(1 for t in txs if t["type"] == "income")
        neutral_count = sum(1 for t in txs if t["type"] == "neutral")
        print(f"[OK] 支出{expense_count} 收入{income_count} 中性{neutral_count} "
              f"→ 新增{added} 跳过{skipped}")

        # Archive
        dest = archive_import(path, source, txs)
        if dest:
            print(f"  已归档: {dest}")

    # Cross-source dedup
    dup_count = find_and_mark_duplicates()
    total_dup += dup_count
    if dup_count > 0:
        print(f"跨源去重: 标记了 {dup_count} 条重复（银行 ↔ 支付平台）")

    # Same-source dedup
    same_dup = dedup_same_source()
    total_dup += same_dup
    if same_dup > 0:
        print(f"同源去重: 标记了 {same_dup} 条重复（相近商户名）")

    if total_dup > 0:
        print(f"总共去重 {total_dup} 条")


def cmd_report(args):
    init_db()
    month = args.month
    if month is None:
        available = get_available_months()
        if not available:
            print("暂无数据，请先导入账单: bill import <文件>", file=sys.stderr)
            sys.exit(1)
        month = available[-1]

    if month not in get_available_months():
        print(f"月份 {month} 无数据", file=sys.stderr)
        sys.exit(1)

    if args.html:
        path = html_report(month, open_browser=not args.no_open)
        if path:
            print(f"HTML 报告已生成: {path}")
    else:
        terminal_report(month)


def cmd_trend(args):
    init_db()
    terminal_trend(args.months)
    # Also show current month category breakdown
    available = get_available_months()
    if available:
        print(f"\n{'═' * 60}")
        print(f"  {available[-1]} 分类支出明细")
        print(f"{'═' * 60}")
        categories = get_category_breakdown(available[-1])
        total_expense = sum(c["expense"] for c in categories)
        if total_expense > 0:
            for cat in categories:
                if cat["expense"] > 0:
                    pct = cat["expense"] / total_expense * 100
                    bar_len = int(pct / 100 * 30)
                    bar = "█" * bar_len
                    print(f"  {cat['category']:<10} {bar} {pct:5.1f}%  ¥{cat['expense']:,.2f}")
        print()


def cmd_query(args):
    init_db()
    txs = query_transactions(month=args.month, include_dup=args.show_dup)
    # Filters
    if args.min:
        txs = [t for t in txs if t["amount"] >= args.min]
    if args.max:
        txs = [t for t in txs if t["amount"] <= args.max]
    if args.type:
        txs = [t for t in txs if t["type"] == args.type]
    if args.source:
        txs = [t for t in txs if t["source"] == args.source]
    if args.category:
        cat = args.category
        txs = [t for t in txs if cat in t.get("category", "")]
    if args.keyword:
        kw = args.keyword.lower()
        txs = [t for t in txs if kw in (t.get("merchant", "") + t.get("product", "") + t.get("remark", "")).lower()]

    if not txs:
        db = get_db()
        cats = db.execute("SELECT DISTINCT category FROM transactions ORDER BY category").fetchall()
        db.close()
        all_cats = sorted(set(c["category"] for c in cats))
        print("无匹配记录", file=sys.stderr)
        if all_cats:
            print(f"可用分类: {', '.join(all_cats)}", file=sys.stderr)
        print("提示: bill query -k <关键词>  或  bill query -c <分类>", file=sys.stderr)
        sys.exit(1)

    # Sort by date desc
    txs.sort(key=lambda x: x["date"], reverse=True)

    plain = args.plain or not sys.stdout.isatty()

    if plain:
        # Machine-readable TSV output
        print("id\tdate\ttype\tcategory\tamount\tmerchant\tsource\tis_duplicate")
        total = 0
        for t in txs:
            type_str = {"expense": "支出", "income": "收入", "neutral": "中性"}.get(t["type"], t["type"])
            print(f"{t['id']}\t{t['date']}\t{type_str}\t{t['category']}\t{t['amount']:.2f}\t{t['merchant']}\t{t['source']}\t{t['is_duplicate']}")
            if t["type"] == "expense":
                total += t["amount"]
        print(f"# {len(txs)} records | expense_total={total:.2f}", file=sys.stderr)
    else:
        w = 100
        print(f"\n{'─' * w}")
        print(f"  {'ID':<6} {'日期':<12} {'类型':<6} {'分类':<10} {'金额':>8} {'商户':<22} {'来源':<8}")
        print(f"{'─' * w}")
        total = 0
        for t in txs:
            type_str = {"expense": "支出", "income": "收入", "neutral": "中性"}.get(t["type"], t["type"])
            dup_mark = " *" if t["is_duplicate"] else ""
            print(f"  {t['id']:<6} {t['date']:<12} {type_str:<6} {t['category']:<10} ¥{t['amount']:>7,.2f} {t['merchant'][:22]:<22} {t['source']:<8}{dup_mark}")
            if t["type"] == "expense":
                total += t["amount"]
        print(f"{'─' * w}")
        print(f"  共 {len(txs)} 笔 | 支出合计: ¥{total:,.2f}")
        print()


def cmd_edit(args):
    init_db()
    db = get_db()
    # Load by ID
    tx = db.execute("SELECT * FROM transactions WHERE id=?", (args.id,)).fetchone()
    if not tx:
        print(f"交易ID {args.id} 不存在", file=sys.stderr)
        db.close()
        sys.exit(1)

    tx = dict(tx)
    if args.category:
        db.execute("UPDATE transactions SET category=? WHERE id=?", (args.category, args.id))
        db.commit()
        print(f"已更新 #{args.id} 分类: {tx['category']} → {args.category}")
        print(f"  {tx['date']} | {tx['merchant']} | ¥{tx['amount']:,.2f}")
    else:
        print(f"#{tx['id']} | {tx['date']} | {tx['merchant']} | ¥{tx['amount']:,.2f} | {tx['category']} | {tx['source']}")
        print("用法: bill edit <ID> --category <新分类>")
    db.close()


def cmd_rules(args):
    if args.add:
        if len(args.add) < 2:
            print("用法: bill rules add <分类> <关键词>")
            return
        add_rule(args.add[0], args.add[1])
    else:
        show_rules()


def cmd_export_guide(args):
    show_export_guide()


def main():
    parser = argparse.ArgumentParser(
        prog="bill",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description="个人月度账单统计工具 — 支持微信/支付宝/中国银行",
        epilog="""使用示例:
  bill import 微信账单.xlsx 支付宝.csv 中行.pdf -p 123456
  bill report                        # 当月终端报告
  bill report 2026-05 --html         # 生成HTML并在浏览器打开
  bill trend -m 12                   # 近12月趋势
  bill query -m 2026-05 -c 餐饮美食   # 检索5月餐饮支出
  bill query -m 2026-05 --min 100    # 5月>=100元的交易
  bill query -k 美团 | sort -t$'\t' -k4 -rn  # 管道到sort排序
  bill edit 42 -c 餐饮美食            # 把ID=42的交易改为餐饮美食
  bill rules                         # 查看分类规则
  bill rules add 餐饮美食 新关键词     # 添加分类关键词
  bill export-guide                  # 各平台账单导出教学
""",
    )
    sub = parser.add_subparsers(dest="command", metavar="<命令>")

    # import
    p_import = sub.add_parser("import", help="导入账单文件（自动识别微信/支付宝/中行）",
        epilog="示例: bill import 微信.xlsx 支付宝.csv 中行.pdf --password 524531")
    p_import.add_argument("files", nargs="+", metavar="文件", help="一个或多个账单文件")
    p_import.add_argument("--password", "-p", help="PDF 密码（中国银行对账单需要）")
    p_import.set_defaults(func=cmd_import)

    # report
    p_report = sub.add_parser("report", help="生成月度收支报告（终端或HTML）",
        epilog="示例: bill report | bill report 2026-05 | bill report --html")
    p_report.add_argument("month", nargs="?", metavar="YYYY-MM",
                          help="指定月份，默认最新有数据的月份")
    p_report.add_argument("--html", action="store_true", help="生成交互式HTML报告（含图表）")
    p_report.add_argument("--no-open", action="store_true", help="不自动用浏览器打开HTML")
    p_report.set_defaults(func=cmd_report)

    # trend
    p_trend = sub.add_parser("trend", help="查看近N月分类支出趋势",
        epilog="示例: bill trend | bill trend -m 12")
    p_trend.add_argument("-m", "--months", type=int, default=6,
                         help="显示最近几个月 (默认: 6)")
    p_trend.set_defaults(func=cmd_trend)

    # query
    p_query = sub.add_parser("query", help="检索交易明细（支持管道输出TSV）",
        epilog="""示例:
  bill query -m 2026-05 -c 餐饮美食
  bill query -m 2026-05 --min 100 --source alipay
  bill query -m 2026-05 | awk -F'\t' 'NR>1{sum+=$4}END{print sum}'
  bill query --keyword 美团""")
    p_query.add_argument("--month", "-m", metavar="YYYY-MM", help="限定月份")
    p_query.add_argument("--min", type=float, metavar="N", help="最低金额")
    p_query.add_argument("--max", type=float, metavar="N", help="最高金额")
    p_query.add_argument("--type", "-t", choices=["expense", "income", "neutral"],
                         help="交易类型: expense/income/neutral")
    p_query.add_argument("--source", "-s", choices=["wechat", "alipay", "boc"],
                         help="来源平台")
    p_query.add_argument("--category", "-c", metavar="分类名", help="交易分类")
    p_query.add_argument("--keyword", "-k", metavar="关键词", help="模糊搜索商户/商品/备注")
    p_query.add_argument("--show-dup", action="store_true", help="包含已去重的重复记录")
    p_query.add_argument("--plain", action="store_true", help="强制TSV输出（管道时自动启用）")
    p_query.set_defaults(func=cmd_query)

    # edit
    p_edit = sub.add_parser("edit", help="修改某笔交易的分类",
        epilog="示例: bill edit 42 -c 餐饮美食")
    p_edit.add_argument("id", type=int, metavar="交易ID", help="交易编号（query 输出第一列）")
    p_edit.add_argument("--category", "-c", metavar="分类名", help="目标分类")
    p_edit.set_defaults(func=cmd_edit)

    # rules
    p_rules = sub.add_parser("rules", help="查看/编辑分类关键词规则",
        epilog="示例: bill rules | bill rules add 餐饮美食 新餐厅名")
    p_rules.add_argument("add", nargs="*", metavar="ARGS",
                         help="添加: bill rules add <分类> <关键词>")
    p_rules.set_defaults(func=cmd_rules)

    # export-guide
    p_guide = sub.add_parser("export-guide", help="显示微信/支付宝/中行的账单导出步骤")
    p_guide.set_defaults(func=cmd_export_guide)

    args = parser.parse_args()
    if args.command is None:
        parser.print_help()
        return
    args.func(args)


if __name__ == "__main__":
    init_db()
    main()
